#!/usr/bin/env python3
"""
deposit_to_excel.py

Convert an FCI Lender Services "Notification of Deposit" PDF into a formatted
Excel workbook. Designed to be invoked from a .bat file via drag-and-drop, but
also runnable directly:

    python deposit_to_excel.py "C:\\path\\to\\Investor_Notification_Deposit.pdf"

The output .xlsx is written next to the input PDF, with the same base name.
"""

import os
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# Dependency check (friendly message if anything is missing)
# ---------------------------------------------------------------------------
_MISSING = []
try:
    import pdfplumber
except ImportError:
    _MISSING.append("pdfplumber")
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule
except ImportError:
    _MISSING.append("openpyxl")

if _MISSING:
    print("ERROR: Missing required Python package(s): " + ", ".join(_MISSING))
    print("Install them with:")
    print("    pip install " + " ".join(_MISSING))
    sys.exit(1)


# ---------------------------------------------------------------------------
# Column layout (in PDF point coordinates) -- determined from the FCI template
# ---------------------------------------------------------------------------
# Each entry: (key, header_text, kind, anchor, x_min, x_max)
#   kind:   "text" | "date" | "int" | "money"
#   anchor: "left"  -> match against word's x0 (left-aligned columns)
#           "right" -> match against word's x1 (right-aligned money columns).
#                     This is what makes wide amounts like "$1,812,590.00" or
#                     "$1,510.49" land in the correct column regardless of how
#                     far left the leading digit sits.
COLUMNS = [
    ("account_no",       "Account No.",                    "text",  "left",    0,    85),
    ("last_name",        "Last Name",                      "text",  "left",   85,   159),
    ("property_address", "Property Address / City, State", "text",  "left",  159,   285),
    ("due_date",         "Payment Due Date",               "date",  "left",  285,   335),
    ("received_date",    "Payment Received",               "date",  "left",  335,   395),
    ("days_late",        "Days Late",                      "int",   "left",  395,   425),
    ("principal_bal",    "Principal Balance",              "money", "right", 425,   522),
    ("principal_pymt",   "Principal Pymt",                 "money", "right", 522,   563),
    ("interest_pymt",    "Interest Pymt",                  "money", "right", 563,   608),
    ("escrow_repymt",    "Escrow Adv Repymt",              "money", "right", 608,   652),
    ("service_fee",      "Service Fee",                    "money", "right", 652,   698),
    ("other_pymts",      "Other Pymts",                    "money", "right", 698,   746),
    ("amount_deposited", "Amount Deposited",               "money", "right", 746,  9999),
]

ACCOUNT_NO_RE = re.compile(r"^\d{9,}$")
MONEY_RE      = re.compile(r"^-?\$[\d,]+\.\d{2}$")
DATE_RE       = re.compile(r"^\d{2}/\d{2}/\d{4}$")
INT_RE        = re.compile(r"^-?\d+$")

Y_LINE_TOLERANCE = 3   # pixels: words within this Y delta are the same physical line


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def money_to_float(s):
    """'$25,347.32' -> 25347.32 ;  '-$15.00' -> -15.00 ;  '' -> None"""
    if s is None or s == "":
        return None
    neg = s.startswith("-")
    s = s.lstrip("-").lstrip("$").replace(",", "")
    try:
        v = float(s)
        return -v if neg else v
    except ValueError:
        return None


def _parse_date(s):
    """'06/09/2026' -> datetime.date(2026, 6, 9); returns None if not parseable."""
    if not s:
        return None
    try:
        return datetime.strptime(s.strip(), "%m/%d/%Y").date()
    except (ValueError, TypeError):
        return None


def _to_int_or_none(s):
    """'4318034' -> 4318034 ; non-digit text -> None"""
    if not s:
        return None
    s = s.strip()
    if INT_RE.match(s):
        try:
            return int(s)
        except ValueError:
            return None
    return None


def _organize_by_account(rows, data_start):
    """Group rows by account_no, preserving first-appearance order.

    Returns:
        ordered_rows: rows reordered so same-account rows are adjacent
        group_spans:  list of (account, first_excel_row, last_excel_row)
    """
    from collections import OrderedDict
    grouped = OrderedDict()
    for r in rows:
        key = r.get("account_no", "")
        grouped.setdefault(key, []).append(r)

    ordered = []
    spans = []
    cursor = data_start
    for acct, group in grouped.items():
        first = cursor
        last = cursor + len(group) - 1
        spans.append((acct, first, last))
        ordered.extend(group)
        cursor = last + 1
    return ordered, spans


def cluster_lines(words, tol=Y_LINE_TOLERANCE):
    """
    Group words into physical lines based on Y position.
    Returns a list of (avg_y, [words sorted by x0]) tuples in top-to-bottom order.
    """
    if not words:
        return []
    words_sorted = sorted(words, key=lambda w: (w["top"], w["x0"]))
    lines = []
    current = [words_sorted[0]]
    current_y = words_sorted[0]["top"]
    for w in words_sorted[1:]:
        if abs(w["top"] - current_y) <= tol:
            current.append(w)
            current_y = (current_y * (len(current) - 1) + w["top"]) / len(current)
        else:
            lines.append((current_y, sorted(current, key=lambda x: x["x0"])))
            current = [w]
            current_y = w["top"]
    lines.append((current_y, sorted(current, key=lambda x: x["x0"])))
    return lines


def assign_columns(line_words):
    """Bucket a line's words into our column ranges.
    Money columns use right-edge (x1) since amounts are right-aligned;
    text/date/int columns use left-edge (x0)."""
    buckets = defaultdict(list)
    for w in line_words:
        for key, _hdr, _kind, anchor, xmin, xmax in COLUMNS:
            x = w["x1"] if anchor == "right" else w["x0"]
            if xmin <= x < xmax:
                buckets[key].append(w["text"])
                break
    return buckets


def join_words(words_list):
    return " ".join(words_list).strip()


# ---------------------------------------------------------------------------
# Summary block parser (top-of-page header block)
# ---------------------------------------------------------------------------
def parse_summary(page):
    """Pull the investor + summary block from the first page."""
    words = page.extract_words(use_text_flow=True, keep_blank_chars=False)
    lines = cluster_lines(words)
    summary = {
        "investor_id":          "",
        "investor_name":        "",
        "date_of_deposit":      "",
        "reference":            "",
        "principal":            None,
        "interest":             None,
        "escrow_adv_repayment": None,
        "servicing_fee":        None,
        "other":                None,
        "other_tax_free":       None,
        "total_deposited":      None,
    }

    # We need to find labels on the right-hand summary block (x ~534) and
    # values on the far right (x ~700+).
    # We also pick up Investor id/name on the left (x ~54..183).
    for _y, lw in lines:
        # left-side Investor block
        texts_left  = [w["text"] for w in lw if w["x0"] < 280]
        texts_right = [w for w in lw if w["x0"] >= 280]
        left_text   = " ".join(texts_left)

        if left_text.startswith("Investor "):
            # "Investor V2401522"
            parts = left_text.split()
            if len(parts) >= 2:
                summary["investor_id"] = parts[1]
        elif left_text and not summary["investor_name"] and summary["investor_id"]:
            # First non-empty left line right after Investor id is the name
            if not any(left_text.startswith(s) for s in ("Account ", "Property ", "City,", "Last ", "Powered ")):
                summary["investor_name"] = left_text

        # Right-side summary key/value pairs.
        # Labels sit roughly between x=534 and x=660; the value sits at x>=660.
        # The word "Summary" appears as a section header around x=480 and must be skipped.
        if not texts_right:
            continue
        label_words = [w["text"] for w in texts_right if 530 <= w["x0"] < 660]
        value_words = [w["text"] for w in texts_right if w["x0"] >= 660]
        label = " ".join(label_words).strip()
        value = " ".join(value_words).strip()
        if label == "Date of Deposit":
            summary["date_of_deposit"] = value
        elif label == "Reference":
            summary["reference"] = value
        elif label == "Principal":
            summary["principal"] = money_to_float(value)
        elif label == "Interest":
            summary["interest"] = money_to_float(value)
        elif label == "Escrow Adv Repayment":
            summary["escrow_adv_repayment"] = money_to_float(value)
        elif label == "Servicing Fee":
            summary["servicing_fee"] = money_to_float(value)
        elif label == "Other":
            summary["other"] = money_to_float(value)
        elif label == "*Other Tax Free":
            summary["other_tax_free"] = money_to_float(value)
        elif label == "Total Deposited":
            summary["total_deposited"] = money_to_float(value)

    return summary


# ---------------------------------------------------------------------------
# Detail rows parser
# ---------------------------------------------------------------------------
def find_table_top_y(lines):
    """Return the Y just below the table header (where data begins)."""
    for y, lw in lines:
        txts = [w["text"] for w in lw]
        if "Account" in txts and "No." in txts:
            # Header may span multiple sub-lines. Data starts a bit further down.
            return y + 5
    return None


def is_footer(words):
    text = " ".join(w["text"] for w in words)
    return text.startswith("Powered by")


def parse_detail_rows(pdf):
    """Return a list of row dicts, one per logical detail row, in document order."""
    rows = []
    totals_row = None  # last-page footer totals row

    for page in pdf.pages:
        words = page.extract_words(use_text_flow=True, keep_blank_chars=False)
        lines = cluster_lines(words)
        table_top = find_table_top_y(lines)
        if table_top is None:
            continue

        # Find lines strictly below the table header that aren't the page footer.
        body_lines = [
            (y, lw) for y, lw in lines
            if y > table_top and not is_footer(lw)
        ]

        current = None
        for y, lw in body_lines:
            buckets = assign_columns(lw)
            acct_cell = " ".join(buckets.get("account_no", [])).strip()

            if ACCOUNT_NO_RE.match(acct_cell):
                # New logical row begins
                if current is not None:
                    rows.append(current)
                current = {key: [] for key, *_ in COLUMNS}
                # Seed with this line's contents
                for k, vs in buckets.items():
                    current[k].extend(vs)
                current["_anchor_y"] = y
            else:
                # Either continuation of previous row, or totals row, or stray text
                # Totals row: no account number, but has values in money columns
                money_cols_present = any(
                    buckets.get(k) for k in
                    ("principal_pymt", "interest_pymt", "escrow_repymt",
                     "service_fee", "other_pymts", "amount_deposited")
                )
                has_name_or_addr = bool(buckets.get("last_name") or buckets.get("property_address"))
                if current is None and money_cols_present and not has_name_or_addr:
                    # Standalone totals row on last page
                    totals_row = buckets
                elif current is not None:
                    # Check vertical distance: continuations are within ~15px of anchor
                    if abs(y - current.get("_anchor_y", y)) <= 18:
                        for k, vs in buckets.items():
                            current[k].extend(vs)
                    else:
                        # Otherwise this is probably the totals row at the end of the page
                        if money_cols_present and not has_name_or_addr:
                            totals_row = buckets
                # else: stray non-data line, skip
        if current is not None:
            rows.append(current)
            current = None

    # Collapse word-lists into single strings
    cleaned = []
    for r in rows:
        cleaned.append({
            key: join_words(r.get(key, []))
            for key, *_ in COLUMNS
        })

    totals_clean = None
    if totals_row:
        totals_clean = {
            key: join_words(totals_row.get(key, []))
            for key, *_ in COLUMNS
        }

    return cleaned, totals_clean


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------
THIN  = Side(style="thin",  color="BBBBBB")
THICK = Side(style="medium", color="000000")

HEADER_FILL = PatternFill("solid", start_color="8B1A2B")  # FCI red-ish
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
LABEL_FONT  = Font(name="Calibri", size=11, bold=True)
TITLE_FONT  = Font(name="Calibri", size=14, bold=True, color="8B1A2B")
NORMAL_FONT = Font(name="Calibri", size=10)
TOTAL_FONT  = Font(name="Calibri", size=10, bold=True)
TOTAL_FILL  = PatternFill("solid", start_color="EDEDED")

MONEY_FMT = '_-$* #,##0.00_-;[Red]-$* #,##0.00_-;_-$* #,##0.00_-;_-@_-'


def write_excel(summary, rows, totals, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Deposit"

    # --- Title row ---
    ws["A1"] = "FCI Lender Services – Notification of Deposit"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:N1")  # spans all data columns including Account Subtotal
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    # --- Summary block ---
    # Coerce summary values into proper Excel types so we don't get
    # "Number stored as text" / "Date stored as text" warning triangles.
    date_of_deposit_val = _parse_date(summary["date_of_deposit"]) or summary["date_of_deposit"] or None
    reference_val       = _to_int_or_none(summary["reference"])  or (summary["reference"] or None)

    sm_rows = [
        ("Investor",              summary["investor_id"]                 or None, "text"),
        ("Investor Name",         summary["investor_name"]               or None, "text"),
        ("Date of Deposit",       date_of_deposit_val,                            "date"),
        ("Reference",             reference_val,                                  "int"),
        ("Principal",             summary["principal"],                           "money"),
        ("Interest",              summary["interest"],                            "money"),
        ("Escrow Adv Repayment",  summary["escrow_adv_repayment"],                "money"),
        ("Servicing Fee",         summary["servicing_fee"],                       "money"),
        ("Other",                 summary["other"],                               "money"),
        ("*Other Tax Free",       summary["other_tax_free"],                      "money"),
        ("Total Deposited",       summary["total_deposited"],                     "money"),
    ]
    start = 3
    for i, (label, value, kind) in enumerate(sm_rows):
        r = start + i
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = LABEL_FONT
        lc.alignment = Alignment(horizontal="left")
        # Value sits in column 3 so it doesn't collide with column-A width
        c = ws.cell(row=r, column=3, value=value)
        c.font = NORMAL_FONT
        c.alignment = Alignment(horizontal="left")
        if kind == "money":
            c.number_format = MONEY_FMT
        elif kind == "date" and value is not None and not isinstance(value, str):
            c.number_format = "mm/dd/yyyy"
        elif kind == "int" and isinstance(value, int):
            c.number_format = "0"  # plain integer, no thousands separator
        if label == "Total Deposited":
            lc.font = Font(name="Calibri", size=11, bold=True, color="8B1A2B")
            c.font  = Font(name="Calibri", size=11, bold=True, color="8B1A2B")

    table_header_row = start + len(sm_rows) + 2  # blank row between summary and table

    # Index of the calculated "Account Subtotal" column (one past the parsed columns)
    SUBTOTAL_COL = len(COLUMNS) + 1
    SUBTOTAL_LETTER = get_column_letter(SUBTOTAL_COL)
    AMOUNT_LETTER   = get_column_letter(len(COLUMNS))  # Amount Deposited

    # --- Column headers (parsed columns + Account Subtotal) ---
    for col_idx, (_key, header, *_rest) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=table_header_row, column=col_idx, value=header)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(left=THIN, right=THIN, top=THICK, bottom=THICK)
    sub_hdr = ws.cell(row=table_header_row, column=SUBTOTAL_COL, value="Account Subtotal")
    sub_hdr.font = HEADER_FONT
    sub_hdr.fill = HEADER_FILL
    sub_hdr.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sub_hdr.border = Border(left=THIN, right=THIN, top=THICK, bottom=THICK)
    ws.row_dimensions[table_header_row].height = 30

    # --- Data rows ---
    data_start = table_header_row + 1

    # Group rows by account so all same-account rows are adjacent.
    ordered_rows, group_spans = _organize_by_account(rows, data_start)

    SUBTOTAL_FILL = PatternFill("solid", start_color="F5F0F0")
    SUBTOTAL_FONT = Font(name="Calibri", size=10, bold=True)

    for i, row in enumerate(ordered_rows):
        excel_row = data_start + i
        for col_idx, (key, _hdr, kind, _anchor, _xmin, _xmax) in enumerate(COLUMNS, start=1):
            raw = row.get(key, "")
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.font = NORMAL_FONT
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            if kind == "money":
                v = money_to_float(raw)
                cell.value = v  # None if blank
                cell.number_format = MONEY_FMT
                cell.alignment = Alignment(horizontal="right")
            elif kind == "int":
                if raw and INT_RE.match(raw):
                    cell.value = int(raw)
                else:
                    cell.value = raw or None
                cell.alignment = Alignment(horizontal="center")
            elif kind == "date":
                # Parse "MM/DD/YYYY" to a real date so Excel doesn't flag it
                # as "date stored as text". Fall back to raw string if parsing fails.
                d = _parse_date(raw)
                if d is not None:
                    cell.value = d
                    cell.number_format = "mm/dd/yyyy"
                else:
                    cell.value = raw or None
                cell.alignment = Alignment(horizontal="center")
            elif key == "account_no":
                # Store account numbers as integers so Excel doesn't show the
                # green "Number Stored as Text" warning triangle.
                iv = _to_int_or_none(raw)
                cell.value = iv if iv is not None else (raw or None)
                cell.number_format = "0"  # no thousands separator
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.value = raw or None
                cell.alignment = Alignment(horizontal="left", wrap_text=True)

        # Empty subtotal cell on every data row (the formula will be added
        # below on the last row of each account group)
        sub_cell = ws.cell(row=excel_row, column=SUBTOTAL_COL)
        sub_cell.font = NORMAL_FONT
        sub_cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        sub_cell.number_format = MONEY_FMT
        sub_cell.alignment = Alignment(horizontal="right")

    data_end = data_start + len(ordered_rows) - 1

    # --- Per-account subtotal formulas (on the last row of each group) ---
    for _acct, first, last in group_spans:
        cell = ws.cell(row=last, column=SUBTOTAL_COL)
        cell.value = f"=SUM({AMOUNT_LETTER}{first}:{AMOUNT_LETTER}{last})"
        cell.font = SUBTOTAL_FONT
        cell.fill = SUBTOTAL_FILL

    # --- Totals row (with formulas) ---
    totals_row = data_end + 1
    ws.cell(row=totals_row, column=1, value="TOTALS").font = TOTAL_FONT
    ws.cell(row=totals_row, column=1).alignment = Alignment(horizontal="right")
    ws.cell(row=totals_row, column=1).fill = TOTAL_FILL
    for col_idx, (_key, _hdr, kind, _anchor, _xmin, _xmax) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=totals_row, column=col_idx)
        c.fill = TOTAL_FILL
        c.font = TOTAL_FONT
        c.border = Border(top=THICK, bottom=THICK, left=THIN, right=THIN)
        if kind == "money" and len(ordered_rows) > 0:
            col_letter = get_column_letter(col_idx)
            c.value = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
            c.number_format = MONEY_FMT
            c.alignment = Alignment(horizontal="right")
    # Grand total for the Account Subtotal column too (sum of per-account subtotals)
    total_sub = ws.cell(row=totals_row, column=SUBTOTAL_COL)
    total_sub.fill = TOTAL_FILL
    total_sub.font = TOTAL_FONT
    total_sub.border = Border(top=THICK, bottom=THICK, left=THIN, right=THIN)
    if len(ordered_rows) > 0:
        total_sub.value = f"=SUM({SUBTOTAL_LETTER}{data_start}:{SUBTOTAL_LETTER}{data_end})"
        total_sub.number_format = MONEY_FMT
        total_sub.alignment = Alignment(horizontal="right")

    # --- Column widths ---
    widths = {
        "account_no":       12,
        "last_name":        36,   # fits "Real Property Multi-Services LLC"
        "property_address": 26,
        "due_date":         13,
        "received_date":    13,
        "days_late":        7,
        "principal_bal":    16,
        "principal_pymt":   14,
        "interest_pymt":    14,
        "escrow_repymt":    15,
        "service_fee":      13,
        "other_pymts":      13,
        "amount_deposited": 15,
    }
    for col_idx, (key, *_rest) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(key, 12)
    ws.column_dimensions[SUBTOTAL_LETTER].width = 16
    # Column A holds both the Account No. data and the summary labels.
    # The summary labels need ~22 chars; make sure A is at least that wide.
    if ws.column_dimensions["A"].width < 22:
        ws.column_dimensions["A"].width = 22

    # NOTE: no freeze_panes here — row 15 (and everything above the table
    # header) is intentionally unlocked so the summary block scrolls
    # together with the data.

    # Highlight negative amounts in red (now includes the subtotal column)
    money_range = f"G{data_start}:{SUBTOTAL_LETTER}{data_end}"
    if len(ordered_rows) > 0:
        ws.conditional_formatting.add(
            money_range,
            CellIsRule(operator="lessThan", formula=["0"],
                       font=Font(name="Calibri", size=10, color="C00000"))
        )

    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def convert(pdf_path, output_dir=None):
    """
    Convert an FCI Notification-of-Deposit PDF into a formatted .xlsx.

    If output_dir is given, the .xlsx is written there (with the PDF's base name).
    Otherwise it's written next to the PDF.
    Returns the Path to the .xlsx that was written.
    """
    pdf_path = Path(pdf_path)
    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF not found: {pdf_path}")
    if output_dir:
        out_dir = Path(output_dir)
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / (pdf_path.stem + ".xlsx")
    else:
        out_path = pdf_path.with_suffix(".xlsx")

    print(f"Reading: {pdf_path}")
    with pdfplumber.open(pdf_path) as pdf:
        print(f"  {len(pdf.pages)} page(s)")
        summary = parse_summary(pdf.pages[0])
        rows, totals = parse_detail_rows(pdf)

    print(f"  Parsed {len(rows)} detail rows")
    print(f"Writing: {out_path}")
    write_excel(summary, rows, totals, out_path)
    print("Done.")
    return out_path


def main():
    if len(sys.argv) < 2:
        print("Usage: python deposit_to_excel.py <input.pdf> [output_dir]")
        print("Or drag and drop a PDF onto the accompanying .bat file.")
        sys.exit(1)
    pdf = sys.argv[1]
    out_dir = sys.argv[2] if len(sys.argv) > 2 else None
    try:
        convert(pdf, output_dir=out_dir)
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback; traceback.print_exc()
        sys.exit(2)


if __name__ == "__main__":
    main()
